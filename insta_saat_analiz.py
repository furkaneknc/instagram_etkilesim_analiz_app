import instaloader
import pandas as pd
from datetime import datetime
import locale

locale.setlocale(locale.LC_ALL,"tr_TR.UTF-8")

#ınstaloadewr kutuphanesınde ornek olustur

L = instaloader.Instaloader()

#Kullanıcı adını al
username = input("Lütfen kullanıcı adını giriniz:")
count = int(input("Kaç Adet Gönderiniz Analiz Edilsin ?"))

#Kullanıcının bilkgilerini indir
profile = instaloader.Profile.from_username(L.context, username)
posts = profile.get_posts()

#analiz icin array olusturucam
post_data =[]

#kullanıcının gönderisini al
for i,post in enumerate(posts):
    if i>= count:
        break
    post_info = {
        "GÜN": post.date.strftime("%A"),
        "AY": post.date.strftime("%B"),
        "YIL": post.date.year,
        "BEĞENİ SAYISI":post.likes,
        "SAAT" : post.date.strftime("%H:%M:%S"),
        "GÖNDERİ LİNKİ": f"https://www.instagram.com/p/{post.shortcode}"
    }

    post_data.append(post_info)


df = pd.DataFrame(post_data)

#Türkçe karakter sorununu ortadan kaldırmak için
df["GÜN"] = df["GÜN"].str.encode("utf-8").str.decode("utf-8")
df["AY"] = df["AY"].str.encode("utf-8").str.decode("utf-8")

#EXCEL dosyasına veri yazma 

excel_file = "instagram_analist.xlsx"
df.to_excel(excel_file,index=False,engine="openpyxl")  

#excel dosyasını açıp, koşullu biçimlendirme yapma
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb= load_workbook(excel_file)
ws = wb.active

#beğeni sayısı 2000 ve üstü olan sütunları yeşile boyayalım.

for row in ws.iter_rows(min_row=2,min_col=4,max_row=len(df)+1,max_col=4):
    for cell in row:
        if cell.value >=2000:
            cell.fill = PatternFill(start_color="00FF00",end_color="00FF00",fill_type="solid")

#excel daosyasını kaydet
wb.save(excel_file)

print(f"Veriler {excel_file} dosyasına yazıldı. Beğeni sayısı 2000 ve üstü olanlar için arka plan yeşile boyandı.")

input("İşlem bitti")